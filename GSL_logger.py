import requests
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ===== CONFIG =====
OWNER = "ahuelsmeier"
REPO = "gsl-transition-generator"
TOKEN = os.getenv("GITHUB_TOKEN")

# Make sure this is your actual Zenodo Record ID
ZENODO_RECORD_ID = "18901586" 

if not TOKEN:
    raise ValueError("Missing GITHUB_TOKEN environment variable")

EXCEL_FILE = "github_traffic_log.xlsx"

GITHUB_BASE_URL = f"https://api.github.com/repos/{OWNER}/{REPO}"
ZENODO_BASE_URL = f"https://zenodo.org/api/records/{ZENODO_RECORD_ID}"

headers_gh = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/vnd.github+json"
}

# ===== FETCH FUNCTIONS =====
def fetch_github_traffic(endpoint):
    url = f"{GITHUB_BASE_URL}/traffic/{endpoint}"
    r = requests.get(url, headers=headers_gh)
    r.raise_for_status()
    return r.json()

def fetch_github_releases():
    """Fetches total download counts for all compiled release assets (.exe, .zip, etc)"""
    url = f"{GITHUB_BASE_URL}/releases"
    r = requests.get(url, headers=headers_gh)
    r.raise_for_status()
    releases = r.json()
    
    total_downloads = 0
    # Loop through all releases and sum up downloads of all attached files
    for release in releases:
        for asset in release.get("assets", []):
            total_downloads += asset.get("download_count", 0)
            
    return total_downloads

def fetch_zenodo():
    """Fetches cumulative stats from Zenodo."""
    r = requests.get(ZENODO_BASE_URL)
    r.raise_for_status()
    data = r.json()
    return data.get("stats", {})

def get_today_summary():
    views = fetch_github_traffic("views")
    clones = fetch_github_traffic("clones")
    releases = fetch_github_releases()
    zenodo = fetch_zenodo()

    today = datetime.utcnow().strftime("%Y-%m-%d")

    return {
        "date": today,
        "views_total": views.get("count", 0),
        "views_unique": views.get("uniques", 0),
        "clones_total": clones.get("count", 0),
        "clones_unique": clones.get("uniques", 0),
        "zenodo_views": zenodo.get("version_views", 0),
        "zenodo_downloads": zenodo.get("version_downloads", 0),
        "release_downloads": releases  # The new binary downloads metric!
    }

# ===== EXCEL HANDLING =====
def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Traffic"

    headers = [
        "date", "views_total", "views_unique", "clones_total", "clones_unique",
        "zenodo_views", "zenodo_downloads", "release_downloads"
    ]
    ws.append(headers)
    wb.save(EXCEL_FILE)

def append_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        init_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Traffic"]

    # --- THE MAGIC UPGRADE TRICK ---
    # Safely add the new column header to your EXISTING file if it's missing
    existing_headers = [cell.value for cell in ws[1]]
    if "release_downloads" not in existing_headers:
        ws.cell(row=1, column=len(existing_headers)+1, value="release_downloads")
    # -------------------------------

    # Avoid duplicate entries for the same day
    dates = [row[0].value for row in ws.iter_rows(min_row=2)]
    if data["date"] in dates:
        print(f"Entry for {data['date']} already exists. Skipping.")
        return

    ws.append([
        data["date"],
        data["views_total"],
        data["views_unique"],
        data["clones_total"],
        data["clones_unique"],
        data["zenodo_views"],
        data["zenodo_downloads"],
        data["release_downloads"]
    ])

    wb.save(EXCEL_FILE)
    print(f"Logged GitHub and Zenodo data for {data['date']}")

# ===== MAIN =====
if __name__ == "__main__":
    data = get_today_summary()
    append_to_excel(data)
